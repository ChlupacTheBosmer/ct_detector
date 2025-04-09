import sqlite3
import pandas as pd
from openpyxl import Workbook


def sqlite_to_excel(db_path, excel_path):
    """
    Convert a SQLite database into an Excel workbook, with each table written to a separate sheet.

    This function connects to a SQLite database specified by 'db_path', retrieves the names of all tables,
    extracts the data from each table into a pandas DataFrame, and then writes each DataFrame to a different
    worksheet in an Excel file defined by 'excel_path'. Uses openpyxl to save the Excel file.

    Parameters:
        db_path (str): The file path to the SQLite database.
        excel_path (str): The file path where the Excel workbook will be saved.

    Returns:
        None

    Raises:
        sqlite3.Error: If an error occurs when connecting to the database or executing SQL commands.
    """
    try:
        # Connect to the SQLite database using the provided file path
        conn = sqlite3.connect(db_path)

        # Create a cursor object to interact with the database
        cursor = conn.cursor()

        # Retrieve the names of all tables present in the database
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
        tables = cursor.fetchall()  # Returns a list of tuples, each containing a table name

        # Create a new workbook using openpyxl
        wb = Workbook()
        # Remove the default sheet created with the workbook (we'll create our own)
        wb.remove(wb.active)

        # Loop through each table name and process its data
        for table in tables:
            table_name = table[0]  # Extract the table name from the tuple
            # Read the entire table into a DataFrame using a SQL query
            df = pd.read_sql_query(f"SELECT * FROM {table_name}", conn)

            # Create a new sheet for the current table
            sheet = wb.create_sheet(title=table_name)

            # Write DataFrame column names as the first row in the sheet
            for col_num, column in enumerate(df.columns, 1):
                sheet.cell(row=1, column=col_num, value=column)

            # Write DataFrame data to the sheet starting from row 2
            for row_num, row in enumerate(df.itertuples(index=False), 2):
                for col_num, value in enumerate(row, 1):
                    sheet.cell(row=row_num, column=col_num, value=value)

        # Save the workbook to the specified Excel path
        wb.save(excel_path)

        # Close the database connection explicitly
        conn.close()

    except sqlite3.Error as e:
        # Output an error message if any SQLite related error occurs
        print("An error occurred while converting SQLite to Excel:", e)


# Example usage: Run the function when executed as the main module.
if __name__ == "__main__":
    # Replace 'example.db' with the path to your SQLite database file
    sqlite_db_path = 'example.db'
    # Replace 'output.xlsx' with the desired path for the output Excel file
    excel_file_path = 'output.xlsx'
    sqlite_to_excel(sqlite_db_path, excel_file_path)
